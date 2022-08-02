# coding: utf-8
import math
from datetime import datetime, timedelta
import os
import glob
# from typing import Any
import openpyxl

# 各種ツイッターのキーをセット
CURRENT_PATH = os.path.join(os.getcwd(), 'excelとpythonで従業員勤怠管理')


def outputExcel(name: str, title: str, head: list, datas):
    '''
    新しくexcelを作成する

    Parameters
    ----------
    name:excel名
    title:シート名
    head:list:シートの1行目
    datas:入れるデータ[[中身],,,]
    '''
    print("Excel----出力開始！----")
    wbOut = openpyxl.Workbook()
    outSheet = wbOut.create_sheet(title=title)
    outSheet.append(head)

    for data in datas:
        # print(outdata)
        outSheet.append(data)

    wbOut.save(os.path.join(CURRENT_PATH, 'output',
                            f'{name}.xlsx'))


def getExcelName(header: str):
    '''
    対象のexcelファイルを取得

    Parameters
    ----------
    header:対象ファイル名(部分一致)

    Returns
    -------
    files:map型
      ID:社員ID
      name: 氏名
      file_path: excelのパス

    '''
    print("Excel----取得開始！----")
    files = []
    # 名前で絞るならこう
    fileName = '*.xlsx'
    inputFiles = glob.glob(os.path.join(CURRENT_PATH, 'input', fileName))
    # 昇順で並び替え(古いやつから処理)
    inputFiles.sort(key=lambda x: x)
    # 条件一致するものの最初のみ取得
    for file in inputFiles:
        if(header in file):
            # print(header + ":::"+file)
            # ファイル名の頭を保持
            # 勤怠管理_ID_名前.xlsx
            title = file.split('/input/')[1]
            title_ID = title.split('_')[1]
            title_name = title.split('_')[2]

            outFile = {
                'ID': title_ID,
                'name': title_name,
                'file_path': os.path.join(CURRENT_PATH, 'input', file),
            }
            files.append(outFile)

    print("Excel取得完了")
    return files


def margeSheetData(sheet):
    '''
    マージ関数→シートデータを配列に

    Parameters
    ----------
    sheet:excelから取得したデータ

    Returns
    -------
    datas:行ごとの2重配列
    maxCount:最終行列数
    '''
    datas = []
    maxCount = 0
    sheet.delete_rows(0)
    maxCount += sheet.max_row

    for row in sheet.rows:
        values = []
        if(row[0].value != None):
            for col in row:
                values.append(col.value)
            datas.append(values)

    return datas, maxCount


def excel_date(num):
    '''
    excelで落ちた数値(日付)を元に戻す
    '''
    return(datetime(1899, 12, 30) + timedelta(days=num))


def sumAttendTime(month, sheet_datas):
    # print(sheet_datas)
    work_hours = 0
    # 該当する日付のみ取得
    # →1月のものに2月のものがある,等
    for data in sheet_datas[0]:
        # print(excel_date(data[0]))
        if(data[4] is not None):
            datetime.datetime.combine(
                datetime.date.today(), basetime) + datetime.timedelta(minutes=1)
            # print(data[4] - data[3])
            print(data[4])
            print(data[3])
            # print(excel_date(data[5]))


def main():
    '''
    流れ
    ・inputでファイルパス等を取得
    ・中身取得
    outFile = openpyxl.load_workbook(
                os.path.join(current_path, 'input', file), data_only=True)
    ・使いやすいように整形
    ・独自処理
    ・outputに出力
    '''

    # print(CURRENT_PATH)
    files = getExcelName("勤怠管理")
    for file in files:
        # print(file)
        outFile = openpyxl.load_workbook(
            file["file_path"], data_only=True)
        # print(outFile.sheetnames)
        # 各シート(日付のみ)で勤怠時間を合計する
        for sheet_name in outFile.sheetnames:
            if(sheet_name != "社員情報"):
                sumAttendTime(sheet_name, margeSheetData(outFile[sheet_name]))


if __name__ == "__main__":
    main()
